from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook


def build_excel_file(
    filename: str, headers: list[str], rows: list[list[object | None]]
) -> tuple[bytes, dict[str, str]]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }


async def read_excel_rows(
    file: UploadFile, expected_headers: list[str]
) -> list[dict[str, object | None]]:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="请上传 .xlsx 文件",
        )

    content = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl exception types vary
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 文件解析失败",
        ) from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 文件不能为空",
        )

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if header_row != expected_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 表头不正确",
        )

    payload: list[dict[str, object | None]] = []
    for row in rows[1:]:
        if row is None:
            continue
        values = list(row)
        if not any(value not in (None, "") for value in values):
            continue
        payload.append(dict(zip(expected_headers, values, strict=False)))
    return payload
